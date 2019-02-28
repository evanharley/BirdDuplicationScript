import openpyxl
import os
import tkinter
from tkinter import filedialog

class BirdDuplicator:

    def __init__(self, *args, **kwargs):
        file_name = filedialog.askopenfilename(title='Select the Item data Spreadsheet',
                                         filetypes=[('xlsx files', '.xlsx')])
        head, tail = os.path.split(file_name)
        self.input_folder = head
        self.filename = tail
        self.output_folder = self.input_folder + '\\new_versions\\'
        self.file = openpyxl.load_workbook(file_name)
        self.ws = self.file['OrnithologyItem']
        self.item_ids = self._create_data_dictionary()
        return super().__init__(*args, **kwargs)

    def set_file_name(file_name):
        head,tail = os.path.split(file_name)
        self.filename = tail

    def parse_spreadsheet(self):
        # takes each of the rows of the spreadsheet and does some logic to inform which helper
        # methods need to be run to handle the row
        last_row = self.ws.max_row
        for sheet in self.file.sheetnames:
            self.ws = self.file[sheet]
            if self.ws[1][1] != 'item_id':
                continue
            for row in range(2, last_row + 1):
                row_data = [self.ws[row][i] for i in range(self.ws.max_column)]
                if row_data[1] in self.item_ids.keys():
                    for item in self.item_ids[row_data[1]]:
                        new_row = row_data
                        new_row[1] = item
                        self.ws.append(new_row)

        self.file.save('{0}\\{1}'.format(self.output_folder, self.filename))
        return 0

    def _create_data_dictionary(self):
        # Takes the Item Spreadsheet and generates a dictionary in the form
        # {item_id: [ids for the component records]}
        data_dictionary = {}
        keys = {self.ws[1][col].value: col for col in range(0, self.ws.max_column)}
        max_id = sorted([self.ws[row][keys['item_id']].value for row in range(2, self.ws.max_row + 1)])[-1]
        for row in range(2, self.ws.max_row + 1):
            item_id = self.ws[row][keys['item_id']].value
            specimen_nature = self.ws[row][keys['specimen_nature']].value.split('; ')
            data_dictionary[item_id] = [max_id + i for i in range(1, len(specimen_nature) + 1)]
            max_id += len(specimen_nature)
                
        return data_dictionary

    def duplicate_birds(self):
        # Main method for duplication of all records
        for dirpath, dirname, filename in os.walk(self.input_folder):
            ignore = ['BIRD', 'RBCM']
            if any(set_file_name.startswith(item) for item in ignore):
                continue
            self.file = openpyxl.load_workbook(filename)
            self.set_file_name(filename)
            self.parse_spreadsheet()

        return 0