import unittest
from BirdDuplicationScript import BirdDuplicator
import openpyxl

class TestBirdDuplication(unittest.TestCase):
    def setUp(self):
        self.dupe = BirdDuplicator()
        self.maxDiff = None
        return super().setUp()

    def test_workbook_exists(self):
        self.assertIsInstance(self.dupe.file, openpyxl.Workbook)

    def test_data_dictionary(self):
        self.dupe._create_data_dictionary()
        test_values = self.dupe.item_ids
        values = {800002:[],
                  816865: [821536, 821537, 821538],
                  820109: [821539, 821540, 821541],
                  821535:[821542, 821543]}
        self.assertEqual(test_values, values)

    def test_component_dictionary(self):
        self.dupe._create_component_dictionary()
        test_values = self.dupe.components
        values = {821536: 'skin, modified', 
                  821537: 'wing', 
                  821538: 'skeleton', 
                  821539: 'skin, partial', 
                  821540: 'wing', 
                  821541: 'skeleton', 
                  821542: 'carcass, partial', 
                  821543: 'wet'}
        self.assertDictEqual(test_values, values)
        self.assertDictEqual(test_values, values)

    def test_parse_spreadsheet(self):
        self.dupe.parse_spreadsheet()
        value_dict = {'Item': [True, []], 
                        'NHItem': [True, []],
                        'OrnithologyItem': [True, []]}
        correct_dict = {'Item': [True, []],
                        'NHItem': [True, []],
                        'OrnithologyItem': [True, []]}
        data = openpyxl.load_workbook('{0}\\{1}'.format(self.dupe.output_folder, self.dupe.filename))
        correct_data = openpyxl.load_workbook('test-correct.xlsx')
        values = []
        for sheet in data.sheetnames:
            ws = data[sheet]
            correct_ws = correct_data[sheet]
            diff = []
            for row in range(2, ws.max_row + 1):
                for col in range(0, ws.max_column):
                    if correct_ws[row][col].value == '':
                        correct_ws[row][col].value = None
                diff.extend([{ws[row][i].value, correct_ws[row][i].value}
                        for i in range(ws.max_column)
                        if ws[row][i].value != correct_ws[row][i].value])
            if diff is not []:
                value_dict[sheet] =  [False, diff]

        self.assertEqual(correct_dict, value_dict)


            



if __name__ == '__main__':
    unittest.main()
